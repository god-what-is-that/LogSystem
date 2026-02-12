import sqlite3
import os
from openpyxl import Workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.utils import get_column_letter
import pandas as pd
from typing import Optional
import warnings

# 高级版本：使用pandas和openpyxl结合，更高效地处理大量数据
def export_logs_to_excel_with_images_advanced(
    db_path: str,
    image_dir: str,
    excel_path: str,
    max_image_width: int = 100,
    max_image_height: int = 100
) -> bool:
    """
    高级版本：先导出数据到Excel，再批量添加图片（性能更好）
    """
    try:
        # 1. 先导出纯数据到Excel（使用pandas快速导出）
        with sqlite3.connect(db_path) as conn:
            query = """
            SELECT 
                id, 
                target, 
                mode, 
                reason, 
                group_id,
                COALESCE(duration, '') as duration,
                COALESCE(operator, '') as operator,
                COALESCE(time, '') as time,
                '' as images_info  -- 预留图片信息列
            FROM logs 
            ORDER BY id DESC
            """
            
            df = pd.read_sql_query(query, conn)
            
            if df.empty:
                print("⚠️ 数据库中没有日志记录")
                return False
        
        # 2. 先保存基础数据到Excel
        print("📊 正在导出数据到Excel...")
        with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='日志记录', index=False)
        
        # 3. 然后打开Excel文件添加图片
        print("🖼️ 正在添加图片...")
        from openpyxl import load_workbook
        wb = load_workbook(excel_path)
        ws = wb['日志记录']
        
        # 确保有图片列标题
        if ws.max_column < 9:
            ws.cell(row=1, column=9, value="images")
        
        # 遍历行，为每条记录添加图片
        for row_idx in range(2, len(df) + 2):  # 从第2行开始（第1行是标题）
            log_id = ws.cell(row=row_idx, column=1).value
            
            if not log_id:
                continue
            
            # 查找图片
            image_col = 9
            image_files_found = []
            
            if os.path.exists(image_dir):
                # 查找以 {id}_ 开头的图片
                for file in os.listdir(image_dir):
                    if file.lower().endswith(('.png', '.jpg', '.jpeg', '.gif', '.bmp', '.webp')):
                        if file.startswith(f"{log_id}_"):
                            image_files_found.append(file)
            
            # 插入图片
            for img_idx, img_file in enumerate(image_files_found):
                img_path = os.path.join(image_dir, img_file)
                
                if os.path.exists(img_path):
                    try:
                        img = ExcelImage(img_path)
                        img.width = max_image_width
                        img.height = max_image_height
                        
                        cell_col = get_column_letter(image_col + img_idx)
                        cell_ref = f"{cell_col}{row_idx}"
                        
                        ws.add_image(img, cell_ref)
                        ws.row_dimensions[row_idx].height = max_image_height * 0.75
                        
                    except Exception as img_error:
                        print(f"  ❌ 添加图片 {img_file} 时出错: {img_error}")
        
        # 4. 调整列宽
        for col in range(1, 9):
            column_letter = get_column_letter(col)
            max_length = 0
            
            for row in range(1, len(df) + 2):
                cell_value = ws.cell(row=row, column=col).value
                if cell_value:
                    max_length = max(max_length, len(str(cell_value)))
            
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # 5. 保存文件
        wb.save(excel_path)
        print(f"✅ Excel文件已保存: {excel_path}")
        print(f"📊 总记录数: {len(df)}")
        
        return True
        
    except Exception as e:
        print(f"❌ 导出失败: {e}")
        import traceback
        traceback.print_exc()
        return False

def start_export():
    # 示例用法
    db_path = "database.db"  # 你的SQLite数据库文件
    image_dir = "./static/images/logs"  # 图片存放目录
    excel_path = "./static/table/logs_with_images.xlsx"  # 生成的Excel文件
    
    # 使用基本版本
    success = export_logs_to_excel_with_images_advanced(
        db_path=db_path,
        image_dir=image_dir,
        excel_path=excel_path,
        max_image_width=120,
        max_image_height=90
    )
    
    return success

# 简单使用示例
if __name__ == "__main__":
    success = start_export()
    
    if success:
        print("🎉 导出完成！")
    else:
        print("😞 导出失败")